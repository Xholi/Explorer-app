# Libraries or packeges we will use 

import os
from numpy import number
from soupsieve import select
import streamlit as st
import pandas as pd  # EDA

from csv import writer

from re import I, M

from this import d

from time import strftime

import pandas as pd

import numpy as np

import datetime

#Visualization
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from numpy import ndarray
from pandas import DataFrame, read_excel
import plotly.express as px
from openpyxl.chart import BarChart, PieChart, Series, Reference

##### For Auto filling the Excel Template if One wants to submit an Excel file####
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import numpy as np

# Create a Work book and Worksheet


########################################################################################################



###########################################################################################################################
def main() :
    """ Sundry/Central Debtors Explorer """
    st.title("Common Debtors Explorer")
    st.subheader("Simple Data Explorer With Streamlit")

    html_temp = """
    <div style="background-color:tomato;"><p style="color:white;font-size:50px;"Steamlit is Awesome</p></div>
    """
    st.markdown(html_temp,unsafe_allow_html=True)

    os.chdir('C:/Users/A245195/python playground/Explorer')
    def file_selector(folder_path = '.\datasets'):
        filenames = os.listdir(folder_path)
        selected_filename = st.selectbox("Select A File",filenames)
        return os.path.join(folder_path,selected_filename)
    filename = file_selector()
    # st.info("You Selected{}".format(filename))

    groupby_column = st.selectbox('What would you like to Analize?',
    ('Ageing','Debit/Credit Ind.')
    )
    # -- Group dataframe
 # Read Data
    df = pd.read_excel(filename, 
    engine='openpyxl',
    sheet_name = 'Data',
    skiprows = 3)
    # df['TAT']= (df['Entry Date']-df['Posting Date'])
    output_columns = ['Debit/Credit Ind.','Amount in local currency','Ageing']
    df_grouped = df.groupby(by=[groupby_column], as_index = False )[output_columns].sum()
    st.title(" Qick Analysis")
    st.dataframe(df_grouped)
    wb = load_workbook('DebtorsTemplate.xlsx')
    ws = wb.active
    
    df['Debit_Credit']=df['Debit/Credit Ind.']
    df['Doc_Type']=df['Document Type']
    df['GL_Account']=df['G/L Account']

    ageing_selection = st.sidebar.multiselect("Select Ageing:",
    options = df["Ageing"].unique(),
    default = df["Ageing"].unique()
)
    Debit_selection = st.sidebar.multiselect("Select Debit/Credit:",
    options = df["Debit_Credit"].unique(),
    default = df["Debit_Credit"].unique()
)
    st.title(" Data Frame")
    df_selection = df.query("Ageing == @ageing_selection & Debit_Credit == @Debit_selection")
    st.dataframe(df_selection)
    if groupby_column == 'Ageing' :
      fig = px.bar(
        df_grouped, x = groupby_column,
        y = 'Amount in local currency',
        color= 'Ageing',
        color_continuous_scale=['red','yellow','green'],
        template = 'plotly_white',
        title= f'<b>Amount in local Currency by {groupby_column}</b>'
    )
    elif groupby_column == 'Debit/Credit Ind.':
     fig = px.bar(
        df_grouped, x = groupby_column,
        y = 'Amount in local currency',
        color= 'Debit/Credit Ind.',
        color_continuous_scale=['red','yellow','green'],
        template = 'plotly_white',
        title= f'<b>Amount in local Currency by {groupby_column}</b>'
    )        
    st.title("Quick Visuals")    
    st.plotly_chart(fig)
    
    if st.button("Column Names"):
        st.write(df.columns)
    # Show Shape
    if st.checkbox("Shape of Dataset"):
        st.write(df.shape)
        data_dim = st.radio("Show Dimension By",("Rows","Columns"))
        if data_dim == "Columns":
            st.text("Number of Columns")
            st.write(df.shape[1])
        if data_dim == "Rows":
            st.text("Number of Rows")
            st.write(df.shape[0])
        else:
            st.write(df.shape)
    # Select Columns
    if st.checkbox("Select Columns To Show"):
        all_columns = df.columns.to_list()
        selected_columns = st.multiselect("Select",all_columns)
        new_df = df[selected_columns]
        st.dataframe(new_df)
#######################################################################################################################
    # Show Summary
    if st.title("Summary of Ageing Analysis"):
        def read_file(data : DataFrame):
            pass


    df_ageing_options : ndarray = df.Ageing.unique()

    def get_summary_by_age(age : str, df : DataFrame):

        debit_filter = df[(df.Ageing == age) & (df["Debit/Credit Ind."] == "Debit")]
        debit_amount_sum = (debit_filter["Amount in local currency"].sum().round(2))
        debit_item_count = len(debit_filter)

        credit_filter = df[(df.Ageing == age) & (df["Debit/Credit Ind."] == "Credit")]
        credit_amount_sum = (credit_filter["Amount in local currency"].sum().round(2))
        credit_item_count = len(credit_filter)

        return DataFrame(
        data=[ [  age, credit_item_count, credit_amount_sum, debit_item_count, debit_amount_sum, (credit_item_count + debit_item_count ), (credit_amount_sum+debit_amount_sum) ] ],
        columns = ["Aging", "Credit Items", "Credit Value", "Debit Items", "Debit Value", "Total Items", "Total Value"])



    data_summary = DataFrame(columns = ["Aging", "Credit Items", "Credit Value", "Debit Items", "Debit Value", "Total Items", "Total Value"])
    for age in df_ageing_options:
        summary_for = get_summary_by_age(age=age, df=df)
        data_summary = data_summary.append(summary_for)
    debit_filter2 = df[(df["Debit/Credit Ind."] == "Debit")]
    debit_filter3 = df[(df["Debit/Credit Ind."] == "Credit")]
    debit_amount_sum = (debit_filter2["Amount in local currency"].sum())
    if st.checkbox("Total Amount"):
        st.write("R",(df["Amount in local currency"].sum().round(2)))
        data_dim = st.radio("Show Amount by",("Debit","Credit"))
        if data_dim == "Credit":
            st.text("Total Credit Amount")
            st.write("R:",(debit_filter3["Amount in local currency"].sum()).round(2))
        if data_dim == "Debit":
            st.text("Total Debit Amount")
            st.write("R:",(debit_filter2["Amount in local currency"].sum()).round(2))

    st.write(data_summary)

#####################################################################################################################################

# ## Plot and Visualization

    # Pie Chart
    if st.checkbox("Pie Plot"):
        all_columns_names = df.columns.to_list()
        if st.button("Generate Pie Plot"):
            st.success("Generating A Pie Plot")
            st.write(df.iloc[:,-1].value_counts().plot.pie(autopct= "%1.1f%%"))
            st.pyplot()

 
    ws['O8'].value = (debit_filter2["Amount in local currency"].sum()).round(2)
    ws['P8'].value = (debit_filter3["Amount in local currency"].sum()).round(2)
    ws['Q8'].value = (df["Amount in local currency"].sum().round(2))


    
    if st.checkbox("Create Excel Template"):
        if st.button("Generate"):
            rows = dataframe_to_rows(data_summary)
            for r_idx, row in enumerate(rows,8):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row = r_idx, column = c_idx,value = value) 
                    date = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")   
        # chart = BarChart()
        # data = Reference(ws, min_row=8, min_col= 1,max_row=15, max_col=12)
        # chart.add_data(data,titles_from_data = True)
        # ws.add_chart(chart,'L12')
        

        
            wb.save(f"DebtorsReport{date}.xlsx")
    

#################################################################################################################################
########################################## THE END OF THE SCRIPT #################################################
st.set_option('deprecation.showPyplotGlobalUse', False)
if __name__ == '__main__' :
    main()
#####################################################################################################################################
# ageing_selection = st.sidebar.multiselect("Select Ageing:",
# options = df["Ageing"].unique(),
# default = df["Ageing"].unique()
# )